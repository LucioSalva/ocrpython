"""XLSX export with 3 sheets: Texto, Campos, Metadatos."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        import json

        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _autosize(ws, max_width: int = 80) -> None:
    for col in ws.columns:
        length = 10
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = max(length, min(max_width, len(str(v).split("\n", 1)[0]) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = length


def export_xlsx(
    *,
    text_content: str | None,
    extracted_fields: dict[str, Any] | None,
    metadata: dict[str, Any] | None,
    document_info: dict[str, Any],
    output_path: Path,
) -> Path:
    wb = Workbook()

    # --- Sheet 1: Texto -----------------------------------------------
    ws_text = wb.active
    ws_text.title = "Texto"
    ws_text["A1"] = "Texto extraído"
    ws_text["A1"].font = Font(bold=True, size=14)
    ws_text.append([])
    body = (text_content or "").splitlines() or [""]
    for line in body:
        ws_text.append([line])
    ws_text.column_dimensions["A"].width = 120

    # --- Sheet 2: Campos ----------------------------------------------
    ws_fields = wb.create_sheet("Campos")
    ws_fields.append(["Campo", "Valor"])
    for cell in ws_fields[1]:
        cell.font = Font(bold=True)
    if extracted_fields:
        for k, v in extracted_fields.items():
            ws_fields.append([k, _stringify(v)])
    else:
        ws_fields.append(["(sin campos extraídos)", ""])
    _autosize(ws_fields)

    # --- Sheet 3: Metadatos -------------------------------------------
    ws_meta = wb.create_sheet("Metadatos")
    ws_meta.append(["Clave", "Valor"])
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)

    rows: list[tuple[str, str]] = [
        ("id", _stringify(document_info.get("id"))),
        ("original_filename", _stringify(document_info.get("original_filename"))),
        ("mime_type", _stringify(document_info.get("mime_type"))),
        ("size_bytes", _stringify(document_info.get("size_bytes"))),
        ("status", _stringify(document_info.get("status"))),
        ("language", _stringify(document_info.get("language"))),
        ("is_native_pdf", _stringify(document_info.get("is_native_pdf"))),
        ("ocr_engine", _stringify(document_info.get("ocr_engine"))),
        ("template_code", _stringify(document_info.get("template_code"))),
        ("created_at", _stringify(document_info.get("created_at"))),
        ("completed_at", _stringify(document_info.get("completed_at"))),
    ]
    for k, v in rows:
        ws_meta.append([k, v])

    if metadata:
        ws_meta.append([])
        ws_meta.append(["-- metadata --", ""])
        for k, v in metadata.items():
            ws_meta.append([str(k), _stringify(v)])

    # Word-wrap on the value column
    for cell in ws_meta["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws_meta)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
